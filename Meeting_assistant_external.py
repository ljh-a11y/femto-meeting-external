import streamlit as st
from streamlit_mic_recorder import mic_recorder
import datetime
import time
import warnings
import os
import openpyxl
from openpyxl.styles import Alignment
import requests
from google_auth_oauthlib.flow import Flow
from google.cloud import storage, speech, firestore
import google.auth
import vertexai
from vertexai.generative_models import GenerativeModel

# --- 시스템 설정 및 경고 제어 ---
warnings.filterwarnings("ignore")

PROJECT_ID = "femto-ai-assistant-497500"
LOCATION = "asia-northeast3"
BUCKET_NAME = "femto-meeting-luke"
TEMPLATE_FILE = "ai 회의록.xlsx"
EXTERNAL_DB_NAME = "external-meetings"

my_options = {"quota_project_id": PROJECT_ID}

# ========================================================
# 🔑 [Streamlit Secrets 보안 키 연동]
# ========================================================
CLIENT_ID = st.secrets["CLIENT_ID"]
CLIENT_SECRET = st.secrets["CLIENT_SECRET"]
REDIRECT_URI = st.secrets["REDIRECT_URI"]


# 구글 OAuth 인증 플로우 생성 함수
def get_google_oauth_flow():
    client_config = {
        "web": {
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "redirect_uris": [REDIRECT_URI]
        }
    }
    flow = Flow.from_client_config(
        client_config,
        scopes=[
            "openid",
            "https://www.googleapis.com/auth/userinfo.email",
            "https://www.googleapis.com/auth/userinfo.profile"
        ],
        redirect_uri=REDIRECT_URI
    )
    flow.autogenerate_code_verifier = False
    return flow


# ========================================================
# [구글 클라우드 백엔드 인증]
# ========================================================
try:
    creds, _ = google.auth.default(quota_project_id=PROJECT_ID)
    vertexai.init(project=PROJECT_ID, location=LOCATION, credentials=creds)
except Exception as e:
    st.error(f"구글 클라우드 인증서 로드 실패: {e}")

# --- 페이지 디자인 ---
st.set_page_config(page_title="Femto AI Dash (External)", layout="wide", initial_sidebar_state="expanded")
st.markdown("""
    <style>
    .stApp { background-color: #F8FAFC; color: #0F172A; }

    iframe[title="streamlit_mic_recorder.mic_recorder"] {
        width: 100% !important;
    }
    </style>
""", unsafe_allow_html=True)

# ========================================================
# 🔑 [사이드바 구글 로그인 & DB 권한 자동 조회]
# ========================================================
st.sidebar.title("🔑 사내 구글 인증")

if "user_email" not in st.session_state:
    st.session_state.user_email = ""
    st.session_state.user_name = ""
    st.session_state.user_level = 0

query_params = st.query_params
if "code" in query_params and not st.session_state.user_email:
    auth_code = query_params["code"]
    try:
        flow = get_google_oauth_flow()
        flow.fetch_token(code=auth_code)
        credentials = flow.credentials

        user_info_response = requests.get(
            "https://www.googleapis.com/oauth2/v2/userinfo",
            headers={"Authorization": f"Bearer {credentials.token}"}
        )
        if user_info_response.status_code == 200:
            google_user = user_info_response.json()
            google_email = google_user.get("email", "").strip().lower()

            db_fs = firestore.Client(project=PROJECT_ID, database=EXTERNAL_DB_NAME, client_options=my_options)
            user_doc = db_fs.collection("users").document(google_email).get()

            if user_doc.exists:
                user_data = user_doc.to_dict()
                if user_data.get("is_active", False):
                    st.session_state.user_email = google_email
                    st.session_state.user_name = user_data.get("name", google_user.get("name", "직원"))
                    st.session_state.user_level = int(user_data.get("level", 1))
                    st.query_params.clear()
                    st.rerun()
                else:
                    st.sidebar.error("❌ 비활성화된 계정입니다.")
            else:
                st.sidebar.error(f"❌ DB에 미등록된 사내 이메일입니다: {google_email}")
    except Exception as err:
        st.sidebar.error(f"구글 인증 연동 실패: {err}")

if not st.session_state.user_email:
    st.sidebar.info("사내 구글 계정으로 로그인해 주세요.")
    flow = get_google_oauth_flow()
    auth_url, _ = flow.authorization_url(prompt='consent')

    st.sidebar.markdown(
        f'''
        <a href="{auth_url}" target="_self" style="text-decoration:none;">
            <div style="background-color: #4285F4; color: white; padding: 12px; border-radius: 8px; text-align: center; font-weight: bold; font-size: 15px;">
                🌐 Google 계정으로 로그인
            </div>
        </a>
        ''',
        unsafe_allow_html=True
    )

if st.session_state.user_email:
    st.sidebar.success("✅ 구글 인증 완료")
    st.sidebar.markdown("---")
    st.sidebar.markdown(f"👤 **접속자:** {st.session_state.user_name}")
    st.sidebar.markdown(f"✉️ **이메일:** {st.session_state.user_email}")
    st.sidebar.markdown(f"🛡️ **권한 등급:** Level {st.session_state.user_level}")
    st.sidebar.markdown("---")

    if st.sidebar.button("로그아웃", use_container_width=True):
        st.session_state.user_email = ""
        st.session_state.user_name = ""
        st.session_state.user_level = 0
        st.query_params.clear()
        st.rerun()


# --- GCS 파일 가져오기 ---
def list_gcs_files():
    try:
        client_s = storage.Client(project=PROJECT_ID, client_options=my_options)
        blobs = list(client_s.list_blobs(BUCKET_NAME))
        return sorted([blob.name for blob in blobs if blob.name.endswith('.wav')], reverse=True)
    except Exception as e:
        st.error(f"구글 스토리지 연결 실패: {e}")
        return []


# --- 1단계: 음성 인식 엔진 ---
def speech_to_text(audio_data, lang_code, gcs_uri=None, selected_filename=None):
    try:
        client_s = storage.Client(project=PROJECT_ID, client_options=my_options)
        bucket = client_s.bucket(BUCKET_NAME)

        if selected_filename:
            txt_filename = selected_filename.replace('.wav', '.txt')
            txt_blob = bucket.blob(txt_filename)
            if txt_blob.exists():
                st.toast("⚡ 저장된 자막 파일(.txt)을 초고속으로 로드했습니다!")
                return txt_blob.download_as_text(encoding="utf-8")

        if not gcs_uri:
            fname = f"meeting_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.wav"
            blob = bucket.blob(fname)
            blob.upload_from_string(audio_data, content_type="audio/wav")
            gcs_uri = f"gs://{BUCKET_NAME}/{fname}"
            selected_filename = fname

        client_speech = speech.SpeechClient(client_options=my_options)
        audio = speech.RecognitionAudio(uri=gcs_uri)

        femto_phrases = [
            "펨토사이언스", "개발혁신팀", "통합사업팀", "제너레이터", "PECVD", "ICP",
            "김무환", "김무환 대표", "김대현", "김대현 고문", "류재홍", "류재홍 박사",
            "김준민", "김준민 팀장", "조인오", "조인오 프로", "윤진성", "윤진성 프로",
            "김다다", "김다다 팀장", "조이정", "조이정 대리", "박홍근", "박홍근 프로", "이주혁", "이주혁 프로"
        ]

        femto_context = speech.SpeechContext(phrases=femto_phrases)

        config = speech.RecognitionConfig(
            language_code=lang_code,
            enable_automatic_punctuation=True,
            use_enhanced=True,
            model="latest_long",
            speech_contexts=[femto_context]
        )

        operation = client_speech.long_running_recognize(config=config, audio=audio)

        response = operation.result(timeout=3600)
        full_text = " ".join([r.alternatives[0].transcript for r in response.results])
        final_result = full_text if full_text.strip() else "(인식된 내용 없음)"

        if final_result and final_result != "(인식된 내용 없음)":
            save_txt_name = selected_filename.replace('.wav', '.txt')
            new_txt_blob = bucket.blob(save_txt_name)
            new_txt_blob.upload_from_string(final_result, content_type="text/plain")

        return final_result
    except Exception as e:
        st.error(f"STT 오류: {e}")
        return ""


# --- 2단계: Gemini 2.5 압축 엔진 ---
def generate_report_text_only(text_content):
    try:
        model = GenerativeModel("gemini-2.5-flash")

        prompt = f"""당신은 '펨토사이언스'의 전문 비서 AI입니다.
        제공된 회의 스크립트를 바탕으로 가시성이 뛰어난 고품질 핵심 회의록 리포트를 작성하세요.

        [ ⚠️ 핵심 제한 조건 - 절대 준수]
        1. 세부 사항을 일일이 나열하지 말고, 자잘한 지시사항은 과감히 통합하거나 생략하여 전체 분량을 콤팩트하게 압축하세요.
        2. 엑셀 출력 시 빈 칸을 포함하여 총 70줄 이내로 끊길 수 있도록 핵심 위주로 요약해야 4페이지를 넘지 않습니다.
        3. 다른 말 없이 무조건 본론 대괄호 헤더부터 즉시 시작하세요.

        [작성 형식]
        - 내용을 더 넓은 범위로 크게 묶어서 최대 4~5개의 [대괄호 주제]로만 분류하세요.
        - 정중하고 깔끔한 비즈니스 개조식 어투(~할 것, ~ 요망, ~ 완료 바람)를 사용하세요.
        - 모든 세부 항목은 개별 줄마다 '• ' 기호로 시작하세요.

        내용: {text_content}"""

        ai_res = model.generate_content(prompt).text
        return ai_res
    except Exception as e:
        st.error(f"리포트 생성 오류: {e}")
        return ""


# --- 3단계: 엑셀 데이터 기입 엔진 ---
def save_final_summary_to_excel(final_summary, meta_info):
    try:
        if os.path.exists(TEMPLATE_FILE):
            wb = openpyxl.load_workbook(TEMPLATE_FILE, data_only=False)
            ws = wb.active

            def safe_write(coord, value, alignment=None):
                target_cell = ws[coord]
                if type(target_cell).__name__ == 'MergedCell':
                    for r_range in ws.merged_cells.ranges:
                        if coord in r_range:
                            target_cell = ws.cell(row=r_range.min_row, column=r_range.min_col)
                            break
                target_cell.value = value
                if alignment:
                    target_cell.alignment = alignment
                return target_cell

            safe_write('B6', meta_info['일시'])
            safe_write('E6', meta_info['부서'])
            safe_write('B7', meta_info['주관'])
            safe_write('E7', meta_info['작성자'])
            safe_write('B8', meta_info['참석자'])
            safe_write('B9', meta_info['안건'])

            p1_rows = list(range(10, 29))
            p2_rows = list(range(32, 59))
            p3_rows = list(range(62, 89))
            p4_rows = list(range(92, 119))

            paragraphs = []
            current_block = []
            category_counter = 1

            for line in final_summary.split('\n'):
                clean_line = line.replace('**', '').replace('###', '').replace('##', '').replace('* ', '• ').strip()
                if not clean_line:
                    continue

                if clean_line.startswith('[') and clean_line.endswith(']'):
                    if current_block:
                        paragraphs.append(current_block)
                    clean_line = f"{category_counter}. {clean_line}"
                    category_counter += 1
                    current_block = [clean_line]
                else:
                    MAX_CHAR_PER_ROW = 60
                    is_bullet = clean_line.startswith('•')
                    indent = "    " if is_bullet else ""

                    temp_text = clean_line
                    is_first_chunk = True

                    while len(temp_text) > MAX_CHAR_PER_ROW:
                        split_idx = temp_text.rfind(' ', 0, MAX_CHAR_PER_ROW)
                        if split_idx == -1 or split_idx == 0:
                            split_idx = MAX_CHAR_PER_ROW

                        chunk = temp_text[:split_idx].strip()
                        if not is_first_chunk:
                            chunk = indent + chunk
                        current_block.append(chunk)
                        temp_text = temp_text[split_idx:].strip()
                        is_first_chunk = False

                    if temp_text:
                        if not is_first_chunk:
                            temp_text = indent + temp_text
                        current_block.append(temp_text)

            if current_block:
                paragraphs.append(current_block)

            final_mapped_lines = []
            p1_idx = p2_idx = p3_idx = p4_idx = 0
            current_page = 1

            for idx_p, block in enumerate(paragraphs):
                block_len = len(block)
                actual_need_len = block_len + (1 if idx_p > 0 else 0)

                if current_page == 1:
                    if p1_idx + actual_need_len <= len(p1_rows):
                        if idx_p > 0:
                            final_mapped_lines.append((p1_rows[p1_idx], ""))
                            p1_idx += 1
                        for b_line in block:
                            final_mapped_lines.append((p1_rows[p1_idx], b_line))
                            p1_idx += 1
                    else:
                        current_page = 2

                if current_page == 2:
                    is_page_start = (p2_idx == 0)
                    need_len_p2 = block_len + (0 if is_page_start else 1)
                    if p2_idx + need_len_p2 <= len(p2_rows):
                        if not is_page_start:
                            final_mapped_lines.append((p2_rows[p2_idx], ""))
                            p2_idx += 1
                        for b_line in block:
                            final_mapped_lines.append((p2_rows[p2_idx], b_line))
                            p2_idx += 1
                    else:
                        current_page = 3

                if current_page == 3:
                    is_page_start = (p3_idx == 0)
                    need_len_p3 = block_len + (0 if is_page_start else 1)
                    if p3_idx + need_len_p3 <= len(p3_rows):
                        if not is_page_start:
                            final_mapped_lines.append((p3_rows[p3_idx], ""))
                            p3_idx += 1
                        for b_line in block:
                            final_mapped_lines.append((p3_rows[p3_idx], b_line))
                            p3_idx += 1
                    else:
                        current_page = 4

                if current_page == 4:
                    is_page_start = (p4_idx == 0)
                    need_len_p4 = block_len + (0 if is_page_start else 1)
                    if p4_idx + need_len_p4 <= len(p4_rows):
                        if not is_page_start:
                            final_mapped_lines.append((p4_rows[p4_idx], ""))
                            p4_idx += 1
                        for b_line in block:
                            final_mapped_lines.append((p4_rows[p4_idx], b_line))
                            p4_idx += 1
                    else:
                        if not is_page_start and p4_idx < len(p4_rows):
                            final_mapped_lines.append((p4_rows[p4_idx], ""))
                            p4_idx += 1
                        for b_line in block:
                            if p4_idx < len(p4_rows):
                                final_mapped_lines.append((p4_rows[p4_idx], b_line))
                                p4_idx += 1

            if current_page == 1:
                if p1_idx < len(p1_rows):
                    final_mapped_lines.append((p1_rows[p1_idx], "-- 이 상 --"))
                else:
                    final_mapped_lines.append((p2_rows[0], "-- 이 상 --"))
                    current_page = 2
            elif current_page == 2:
                if p2_idx < len(p2_rows):
                    final_mapped_lines.append((p2_rows[p2_idx], "-- 이 상 --"))
                else:
                    final_mapped_lines.append((p3_rows[0], "-- 이 상 --"))
                    current_page = 3
            elif current_page == 3:
                if p3_idx < len(p3_rows):
                    final_mapped_lines.append((p3_rows[p3_idx], "-- 이 상 --"))
                else:
                    final_mapped_lines.append((p4_rows[0], "-- 이 상 --"))
                    current_page = 4
            elif current_page == 4:
                if p4_idx < len(p4_rows):
                    final_mapped_lines.append((p4_rows[p4_idx], "-- 이 상 --"))
                else:
                    final_mapped_lines.append((p4_rows[-1], "-- 이 상 --"))

            used_rows = [item[0] for item in final_mapped_lines]
            max_used_row = max(used_rows) if used_rows else 10

            if max_used_row <= 28:
                total_pages = 1
            elif max_used_row <= 58:
                total_pages = 2
            elif max_used_row <= 88:
                total_pages = 3
            else:
                total_pages = 4

            safe_write('A4', f"회  의  록 ( 1 / {total_pages} )")
            if total_pages >= 2:
                safe_write('A30', f"회  의  록 ( 2 / {total_pages} )")
            if total_pages >= 3:
                safe_write('A60', f"회  의  록 ( 3 / {total_pages} )")
            if total_pages >= 4:
                safe_write('A90', f"회  의  록 ( 4 / {total_pages} )")

            if total_pages == 1:
                for r_clear in range(29, 130):
                    ws.row_dimensions[r_clear].hidden = True
            elif total_pages == 2:
                for r_clear in range(59, 130):
                    ws.row_dimensions[r_clear].hidden = True
            elif total_pages == 3:
                for r_clear in range(89, 130):
                    ws.row_dimensions[r_clear].hidden = True
            elif total_pages == 4:
                for r_clear in range(119, 130):
                    ws.row_dimensions[r_clear].hidden = True

            for r, line in final_mapped_lines:
                ws.row_dimensions[r].height = 22
                coord = f'B{r}'
                if line == "-- 이 상 --":
                    written_cell = safe_write(coord, line)
                    written_cell.alignment = Alignment(vertical='center', horizontal='center')
                else:
                    written_cell = safe_write(coord, line)
                    written_cell.alignment = Alignment(vertical='center', horizontal='left')

            output_filename = f"회의록_{datetime.date.today()}_{meta_info['주관']}.xlsx"
            wb.save(output_filename)
            return output_filename
        else:
            st.error("폴더 내에 'ai 회의록.xlsx' 원본 템플릿 파일이 없습니다.")
            return None
    except Exception as e:
        st.error(f"엑셀 빌드 오류: {e}")
        return None


# --- 메인 대시보드 UI ---
st.markdown("""
    <div style="background-color: #0F172A; padding: 20px; border-radius: 16px; margin-bottom: 30px; border-left: 6px solid #3B82F6;">
        <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 800; letter-spacing: -0.5px;">🏢 FEMTO SCIENCE AI COMMAND CENTER (EXTERNAL)</h1>
        <p style="color: #94A3B8; margin: 5px 0 0 0; font-size: 13px; font-weight: 600;">EXTERNAL SERVICE ARCHIVE &amp; LIVE CONSOLE</p>
    </div>
""", unsafe_allow_html=True)

tab1, tab2 = st.tabs(["🚀 LIVE CONSOLE", "📁 MEETING ARCHIVE"])

with tab1:
    col_meta, col_action = st.columns([1.1, 1], gap="large")

    with col_meta:
        st.subheader("📋 회의 기본 정보 입력")

        dept_options = ["통합사업팀 (IBD Team)", "개발혁신팀", "직접입력"]
        selected_dept = st.selectbox("부서 선택", dept_options, index=0)
        m_dept = st.text_input("부서명 수기 입력", placeholder="부서명을 직접 입력하세요") if selected_dept == "직접입력" else selected_dept

        author_options = [
            "이주혁 프로", "박홍근 프로", "조이정 대리", "김다다 팀장",
            "윤진성 프로", "조인오 프로", "김준민 팀장", "김무환 대표",
            "김대현 고문", "류재홍 박사", "직접입력"
        ]
        selected_author = st.selectbox("작성자 선택", author_options, index=0)
        m_writer = st.text_input("작성자 수기 입력",
                                 placeholder="이름 및 직급을 입력하세요") if selected_author == "직접입력" else selected_author

        m_host = st.text_input("주관 (발표자/회의 리더)", "이주혁")

        attendee_options = ["전직원(대표님 포함)", "통합사업팀", "개발혁신팀", "수기입력"]
        selected_attendees = st.selectbox("참석자 범위 선택", attendee_options, index=0)
        m_attendees = st.text_input("참석자 명단 수기 입력",
                                    placeholder="참석자 명단을 입력하세요") if selected_attendees == "수기입력" else selected_attendees

        topic_options = [f"정기 업무 보고_{datetime.date.today()}", "월발회의", "팀장회의", "팀 회의", "수기입력"]
        selected_topic = st.selectbox("안건 종류 선택", topic_options, index=0)
        m_title = st.text_input("안건 수기 입력",
                                placeholder="회의 제목이나 안건을 입력하세요") if selected_topic == "수기입력" else selected_topic

        m_date = st.text_input("일시", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))

        meta_data = {"부서": m_dept, "작성자": m_writer, "주관": m_host, "참석자": m_attendees, "안건": m_title, "일시": m_date}

    with col_action:
        st.subheader("🛡️ AI ANALYSIS TIER (보안 레벨)")
        security_level = st.select_slider(
            "회의 보안 등급 설정",
            options=["Level 1 (Public)", "Level 2 (Internal)", "Level 3 (Secret)"],
            help="설정된 보안 레벨에 따라 해당 등급 이상의 권한을 가진 사용자만 열람할 수 있습니다."
        )
        st.write(f"🔒 활성화된 보안 상태: **{security_level}**")
        st.divider()

        st.subheader("🎙️ 오디오 제어 / 회의록 생성 모듈")
        input_mode = st.radio("작업 방식 선택", ["실시간 녹음", "기존 파일 선택", "📝 회의록 텍스트 직접 입력"])

        if "is_recording" not in st.session_state:
            st.session_state.is_recording = False
        if "is_converting" not in st.session_state:
            st.session_state.is_converting = False

        if input_mode == "실시간 녹음":
            st.markdown("""
                <div style='background-color: #EFF6FF; border: 1px solid #BFDBFE; padding: 14px 18px; border-radius: 12px; margin-bottom: 12px;'>
                    <span style='font-size: 14px; font-weight: 700; color: #1E40AF;'>🎙️ 실시간 음성 녹음 안내</span><br>
                    <span style='font-size: 12px; color: #475569;'>아래 <b>[🔴 실시간 녹음 시작]</b> 버튼을 누르면 마이크가 켜집니다. 회의 종료 후 한 번 더 누르면 AI 회의록이 생성됩니다.</span>
                </div>
            """, unsafe_allow_html=True)

            audio_record = mic_recorder(
                start_prompt="🔴 실시간 녹음 시작 (마이크 활성화)",
                stop_prompt="⏹️ 녹음 완료 및 AI 회의록 즉시 생성",
                key='rec'
            )

            if st.session_state.is_converting:
                with st.status("🔮 AI 음성 분석 및 회의록 생성 중...", expanded=True) as status:
                    st.write("1️⃣ 오디오 데이터 수신 완료")
                    time.sleep(0.8)
                    st.write("2️⃣ 구글 STT 자막 변환 처리 중...")
                    time.sleep(0.8)
                    status.update(label="✅ 회의록 생성 완료!", state="complete", expanded=False)

            if audio_record:
                st.session_state.is_recording = False
                st.session_state.is_converting = True
                st.session_state.transcript = speech_to_text(audio_record['bytes'], "ko-KR")
                st.session_state.is_converting = False
                st.rerun()

        elif input_mode == "기존 파일 선택":
            gcs_files = list_gcs_files()
            if gcs_files:
                selected_file = st.selectbox("클라우드 오디오 파일 선택", gcs_files)
                if st.button("🚀 선택 파일 AI 회의록 생성 시작", type="primary", use_container_width=True):
                    st.session_state.is_converting = True
                    gcs_uri = f"gs://{BUCKET_NAME}/{selected_file}"
                    st.session_state.transcript = speech_to_text(None, "ko-KR", gcs_uri=gcs_uri,
                                                                 selected_filename=selected_file)
                    st.session_state.is_converting = False
                    st.rerun()
            else:
                st.warning("구글 스토리지에서 오디오(.wav) 목록을 불러오지 못했거나 파일이 없습니다.")

        elif input_mode == "📝 회의록 텍스트 직접 입력":
            st.info("💡 텍스트 입력 후 엑셀 양식 단독 빌드 테스트를 수행합니다.")
            sandbox_text = st.text_area("엑셀 반영 테스트용 내용 입력", height=200)

            if st.button("📊 테스트 엑셀 문서 즉시 빌드", type="primary", use_container_width=True):
                st.session_state.ai_summary = sandbox_text
                st.session_state.transcript = "(엑셀 양식 단독 레이아웃 테스트 모드로 생성됨)"
                with st.spinner("템플릿 양식에 맞춰 엑셀 빌드 중..."):
                    excel_path = save_final_summary_to_excel(sandbox_text, meta_data)
                    if excel_path:
                        st.session_state.excel_path = excel_path

    if input_mode != "📝 회의록 텍스트 직접 입력" and 'transcript' in st.session_state:
        st.divider()
        st.subheader("📝 1차 음성 인식 결과")
        edited_text = st.text_area("인식된 대화 내용 (STT)", value=st.session_state.transcript, height=220)

        if st.button("✨ 1차 AI 요약 리포트 초안 생성", type="secondary", use_container_width=True):
            st.session_state.transcript = edited_text
            with st.spinner("Gemini가 회의 내용을 요약하고 있습니다..."):
                ai_summary = generate_report_text_only(edited_text)
                if ai_summary:
                    st.session_state.ai_summary = ai_summary
                    if 'excel_path' in st.session_state:
                        del st.session_state['excel_path']

    if input_mode != "📝 회의록 텍스트 직접 입력" and 'ai_summary' in st.session_state:
        st.divider()
        st.subheader("🤖 AI 최종 요약 리포트 검토")
        edited_summary = st.text_area("엑셀에 반영될 요약 내용 편집", value=st.session_state.ai_summary, height=300)

        if st.button("📊 이 내용으로 최종 엑셀 문서 빌드", type="primary", use_container_width=True):
            st.session_state.ai_summary = edited_summary
            with st.spinner("편집된 내용을 토대로 'ai 회의록.xlsx' 템플릿에 쓰는 중..."):
                excel_path = save_final_summary_to_excel(edited_summary, meta_data)
                if excel_path:
                    st.session_state.excel_path = excel_path

                    try:
                        db = firestore.Client(project=PROJECT_ID, database=EXTERNAL_DB_NAME, client_options=my_options)
                        db.collection("meetings").add({
                            "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "title": meta_data['안건'],
                            "host": meta_data['주관'],
                            "dept": meta_data['부서'],
                            "writer": meta_data['작성자'],
                            "writer_email": st.session_state.user_email,
                            "attendees": meta_data['참석자'],
                            "transcript": st.session_state.transcript,
                            "ai_summary": st.session_state.ai_summary,
                            "security_level": security_level
                        })
                        st.toast("🔥 외부 전용 클라우드 DB에 회의록 기록이 백업되었습니다!")
                    except Exception as db_err:
                        st.warning(f"DB 백업 중 권한 오류 발생: {db_err}")

    if 'excel_path' in st.session_state:
        st.divider()
        st.success("🎉 최종 엑셀 회의록 문서 생성이 완료되었습니다!")

        with open(st.session_state.excel_path, "rb") as f:
            st.download_button(
                label="📥 완성된 엑셀 회의록 다운로드",
                data=f,
                file_name=st.session_state.excel_path,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

# ========================================================
# 📂 [Tab 2] Meeting Archive (로그인 권한별 열람 제한)
# ========================================================
with tab2:
    st.subheader("📂 Meeting Archive (외부 전용 과거 이력 보관소)")

    if not st.session_state.user_email:
        st.warning("⚠️ 사이드바에서 구글 로그인을 먼저 진행하셔야 회의록을 열람하실 수 있습니다.")
    else:
        st.caption(
            f"접속자: **{st.session_state.user_name}** ({st.session_state.user_email}) | 열람 권한: **Level {st.session_state.user_level} 이하 전체**")

        allowed_levels = []
        if st.session_state.user_level >= 1:
            allowed_levels.append("Level 1 (Public)")
        if st.session_state.user_level >= 2:
            allowed_levels.append("Level 2 (Internal)")
        if st.session_state.user_level >= 3:
            allowed_levels.append("Level 3 (Secret)")

        try:
            db = firestore.Client(project=PROJECT_ID, database=EXTERNAL_DB_NAME, client_options=my_options)

            docs = list(
                db.collection("meetings")
                .where("security_level", "in", allowed_levels)
                .order_by("date", direction=firestore.Query.DESCENDING)
                .stream()
            )

            if not docs:
                st.info("열람 가능한 회의록 이력이 없거나 접근 권한 범위를 벗어납니다.")

            for doc in docs:
                data = doc.to_dict()
                doc_id = doc.id

                with st.expander(
                        f"📅 {data.get('date', '날짜 정보 없음')} | 🏢 {data.get('dept', '부서 미지정')} | 📋 {data.get('title', '제목 없음')} ({data.get('security_level', 'Level 1')})"):
                    st.markdown(
                        f"**🗣️ 회의 주관:** {data.get('host', '-')}   |   **✍️ 작성자:** {data.get('writer', '-')}   |   **👥 참석자:** {data.get('attendees', '-')}")
                    st.markdown("---")

                    st.markdown("### 🤖 AI 최종 요약 리포트 본문")
                    st.info(data.get('ai_summary', '(요약 내용 없음)'))

                    c1, c2 = st.columns([1, 4])
                    with c1:
                        is_author = (data.get('writer_email') == st.session_state.user_email)
                        if st.session_state.user_level == 3 or is_author:
                            if st.button(f"🗑️ 기록 삭제", key=f"del_{doc_id}", help="클라우드 DB에서 이 회의록을 영구히 지웁니다."):
                                db.collection("meetings").document(doc_id).delete()
                                st.toast("선택하신 회의록 이력이 정상적으로 삭제되었습니다.")
                                time.sleep(0.8)
                                st.rerun()
                    with c2:
                        with st.expander("🔍 회의 원본 스크립트(STT) 전문 보기"):
                            st.write(data.get('transcript', '(인식된 원본 텍스트가 없습니다.)'))

        except Exception as e:
            st.error(f"데이터베이스 조회 오류: {e}")